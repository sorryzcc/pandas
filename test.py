import openpyxl
from openpyxl.utils import get_column_letter

# 创建一个新的工作簿和工作表
wb = openpyxl.Workbook()
ws = wb.active

# 给工作表添加一些数据作为示例
for row in range(1, 10):  # 添加9行数据
    for col in range(1, 5):  # 每行4列
        cell = ws.cell(row=row+1, column=col)
        cell.value = f"Row {row}, Col {col}"

# 设置首行的标题
for col_num, header_title in enumerate(['Header1', 'Header2', 'Header3', 'Header4'], start=1):
    ws.cell(row=1, column=col_num).value = header_title

# 冻结首行
ws.freeze_panes = 'A2'  # 这将冻结第一行

# 保存工作簿到文件
wb.save('fixed_header_example.xlsx')